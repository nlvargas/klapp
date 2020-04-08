from collections import OrderedDict
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import json
import ast

class Student:
    def __init__(self, id, attributes, preferences, disponibilities):
        self.id = id
        self.attributes = attributes
        self.preferences = preferences
        self.answered = True if preferences != ['None']*len(preferences) else False
        self.disponibilities = disponibilities
        self.priority = defaultdict(lambda: 100)
        self.flexibility = 1
        self.a = {}

    def __repr__(self):
        return "{}: Atributos: {} - Preferencias: {} - Disponibilidad: {}"\
               .format(self.id, self.attributes, self.preferences, self.a)

    def __str__(self):
        return self.__repr__()

    def get_disponibility(self, disponibilities_name):
        for i, j in zip(disponibilities_name, self.disponibilities):
            print(i, j)
            if j == 1:
                self.flexibility += 1
            self.a[i] = int(j)

    def get_priorities(self, groups):
        for i, preference in enumerate(self.preferences):
            groups_prefered = list(filter(lambda g: "Grupo {} - ".format(preference) in g, groups))
            for g in groups_prefered:
                self.priority[g] = i + 1


def get_students_priorities(students, groups):
    for student in students:
        s = students[student]
        s.get_priorities(groups)


def create_students(attributes, disponibilities, students_preferences_number, params):
    students = {}
    for p in params:
        s = Student(p["ID"],
                    [p[a] for a in attributes],
                    [p[pref] for pref in range(1, students_preferences_number + 1)],
                    [p[d] for d in disponibilities])
        if disponibilities:
            s.get_disponibility(disponibilities)
        #print(s.id, s.a)
        s.get_priorities(s.preferences)
        #print(s.id, s.priority)
        students[s.id] = s
    return students


def excel_to_dict(params):
    attributes = params["attributes"]
    disponibilities = params["disponibilities"] if "disponibilities" in params else []
    students_preferences_number = params["students_preferences_number"]
    wb = load_workbook('groups/groups_generator/input.xlsx')
    ws = wb["Alumnos"]
    l = []
    parameters = ["ID"] + \
                 [a for a in attributes] + \
                 [d for d in disponibilities] + \
                 [p for p in range(1, students_preferences_number + 1)]
    for row in ws.iter_rows(min_row=2, min_col=0, max_col=len(parameters)):
        d = OrderedDict()
        null_count = 0
        for i in range(0, len(parameters)):
            d[parameters[i]] = str(row[i].value)
            if row[i].value is None:
                null_count += 1
        if null_count >= len(parameters):
            break
        l.append(d)
    return l


def student_attributes(attributes):
    wb = load_workbook('groups/groups_generator/input.xlsx')
    ws = wb["Alumnos"]
    dic_alumnos = {}
    for row in ws.iter_rows(min_row=2, min_col=0, max_col=len(attributes) + 2):
        d = defaultdict(lambda: 0)
        for i in range(1, len(attributes) + 1):
            d[str(row[i].value)] = 1
        dic_alumnos[str(row[0].value)] = d
    return dic_alumnos


def groups_prefered(student, G):
    if student.preferences == ['None'] * len(student.preferences):
        return G
    else:
        gr = []
        for preference in student.preferences:
            groups = list(filter(lambda g: "Grupo {} - ".format(preference) in g, G))
            gr = gr + groups
        if gr:
            return gr
        else:
            return G


def create_template(params):
    students_preferences_number = params["students_preferences_number"]
    attributes = params["attributes"]
    disponibilities = params["divisions"]
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Alumnos"
    r = ["Nombre"] + \
        [a for a in attributes] + \
        ["Disponibilidad {}".format(d) for d in disponibilities] + \
        ["Preferencia {}".format(i) for i in range(1, students_preferences_number + 1)]
    sheet.append(r)
    sheet2 = wb.create_sheet("Opciones Adicionales")
    sheet2.append(["", "Min por grupo", "Max por grupo", "Pueden haber grupos sin el atributo?"])

    wb.save('groups/groups_generator/template.xlsx')


def get_attributes_bounds(upper_number):
    wb = load_workbook('groups/groups_generator/input.xlsx')
    ws = wb["Opciones Adicionales"]
    d = {}
    for row in ws.iter_rows(min_row=2, min_col=0, max_col=5, max_row=1000000):
        if row[0].value is None:
            break
        d_options = {}
        d_options["min"] = int(row[1].value) if row[1].value is not None else 0
        d_options["max"] = int(row[2].value) if row[2].value is not None else upper_number
        if row[3].value is not None:
            d_options["solo"] = True
        else:
            d_options["solo"] = False
        d[row[0].value] = d_options
    return d


def save_results(I, G, y, students):
    res = []
    for g in G:
        gr = [i for i in I if y[i, g].X != 0]
        if len(gr) > 1:
            res.append([])
            res.append([g])
        for alumno in gr:
            res.append([alumno] + students[alumno].preferences + [a for a in students[alumno].attributes])
    # print(res)
    book = Workbook()
    sheet = book.active
    for row in res:
        if type(row) == type([1]):
            sheet.append(row)
        else:
            sheet.append([row])
    book.save('groups/groups_generator/results.xlsx')


def clean_dict(query_dict):
    dic = {}
    cap = {}
    prefs = {}
    for element in query_dict:
        if element[-2:] == "[]":
            dic[element[:-2]] = query_dict.getlist(element)
        elif "capacity" in element:
            name = element[9:-1]
            cap[name] = int(query_dict[element])
        elif "preferences" in element and "students_preferences_number" not in element:
            name = element[12:-6]
            if name not in prefs:
                prefs[name] = {}
            if "min" in element:
                prefs[name]["min"] = int(query_dict[element])
            else:
                prefs[name]["max"] = int(query_dict[element])

            #prefs = json.loads(query_dict[element])
            #print(prefs)
            #for pre in prefs:
            #    prefs[pre]["min"] = int(prefs[pre]["min"])
            #    prefs[pre]["max"] = int(prefs[pre]["max"])

        else:
            dic[element] = int(query_dict[element])
    dic["capacity"] = cap
    dic["preferences"] = prefs
    return dic


def count_attributes(students):
    attributes = {}
    for id in students:
        for a in students[id].attributes:
            if a not in attributes:
                attributes[a] = 1
            else:
                attributes[a] += 1
    return attributes









