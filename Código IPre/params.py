from collections import OrderedDict
from openpyxl import load_workbook

class Student:
    def __init__(self, id, preferences, gender=None, area=None,
                 location=None, career=None, university=None, type=None):
        self.id = id
        self.gender = gender
        self.area = area
        self.location = location
        self.career = career
        self.university = university
        self.type = type
        self.preferences = preferences
        self.priority = 1000000000
        self.flexibility = 100000000
        self.asignation = None

    def __repr__(self):
        return str(self.id)

    def __str__(self):
        return self.__repr__()


def create_students(params):
    students = []
    for p in params:
        s = Student(p["ALUMNO"], [p["P1"], p["P2"], p["P3"]], p["GENERO"],
                    p["AREA"], p["COMUNA"], p["CARRERA"], p["UNI"], p["TIPO"])
        students.append(s)
    return students



def excel_to_dict(filename, worksheet_name, parameters, first_row, first_col, last_col):
    wb = load_workbook(filename)
    ws = wb[worksheet_name]
    list = []

    for row in ws.iter_rows(min_row=first_row, max_col=last_col, min_col=first_col):
        dict = OrderedDict()
        null_count = 0

        for i in range(0, len(parameters)):
            dict[parameters[i]] = str(row[i].value)
            if row[i].value is None:
                null_count += 1

        if null_count >= len(parameters):
            break
        list.append(dict)
    return list



vars = ["ALUMNO", "GENERO", "AREA", "COMUNA", "CARRERA", "UNI", "TIPO", "P1", "P2", "P3"]
params = excel_to_dict('Caracterizacion_Estudiantes_MBA_2019-1.xlsm', "Hoja1", vars, 2, 0, 10)
students = create_students(params)


